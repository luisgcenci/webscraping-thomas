from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time
import json
from openpyxl import Workbook, load_workbook

PATH = "./chromedriver.exe"

def create_usssa_team(team_name, team_class, team_age, team_city, team_state, team_coach, usssa_db):
    usssa_db[team_name] = {
        'team_name': team_name, 
        'team_city': team_city,
        'team_state': team_state, 
        'team_age': team_age, 
        'team_class':team_class, 
        'team_points': "",
        'team_coach' : "",
        'tournaments': 
        {
            'FASA': 
            { 
                'tournaments': {}, 
                'games': {},
                'upcoming_tournaments': {}
            }, 
            'USSSA': 
            { 
                'tournaments': {}, 
                'games': {},
                'upcoming_tournaments': {}
            }, 
            'USFA': 
            { 
                'tournaments': {}, 
                'games': {},
                'upcoming_tournaments': {}
            }
        }
    }

def create_team(team_name, team_class, team_age, team_city, team_state, team_coach, teams_db):
    
    teams_db[team_name] = {
        'team_name': team_name, 
        'team_city': team_city,
        'team_state': team_state, 
        'team_age': team_age, 
        'team_points': "", 
        'fasa_class':"",
        'usfa_class':"",
        'usssa_class':team_class,
        'team_coach' : team_coach,
        'tournaments':
        {
            'FASA': 
            { 
                'tournaments': {}, 
                'games': {},
                'upcoming_tournaments': {}
            }, 
            'USSSA': 
            { 
                'tournaments': {}, 
                'games': {},
                'upcoming_tournaments': {}
            }, 
            'USFA': 
            { 
                'tournaments': {}, 
                'games': {},
                'upcoming_tournaments': {}
            }
        }
    }

def get_results(team_1_score, team_2_score):
    
    team_1_result = ""
    team_2_result = ""

    if team_1_score == None or team_2_score == None or team_1_score == team_2_score:
        team_1_result = "T"
        team_2_result = "T"
    
    elif team_1_score > team_2_score:
        team_1_result = "W"
        team_2_result = "L"
    
    elif team_1_score < team_2_score:
        team_1_result = "L"
        team_2_result = "W"
    
    return (team_1_result, team_2_result)

def getTeamInfo(usssa_db, teams_db, team_class, team_url, game_id):

    driver = webdriver.Chrome(PATH)
    driver.get(team_url)

    teams = driver.find_elements_by_css_selector('.table-responsive-sm > table > tbody > tr')
    len_teams = len(teams)
    count = 0
    for i in range(len_teams):
        driver.get(team_url)

        try:
            team = driver.find_elements_by_css_selector('.table-responsive-sm > table > tbody > tr').pop(i)
            team_name = team.find_element_by_tag_name('td').text.strip()
            team_link = team.find_element_by_css_selector('td > a')
            
            team_info = team.find_elements_by_tag_name('td')
            team_age = team_info.pop(1).text.strip()
            team_loc = team_info.pop(1).text.strip().split(',')
            team_city = team_loc[0].strip()
            team_state = team_loc[1].strip()
            team_coach = team_info.pop(1).text.strip()

        except NoSuchElementException or IndexError:
            break
        
        if team_name not in usssa_db.keys():
            create_usssa_team(team_name, team_class, team_age, team_city, team_state, team_coach,  usssa_db)
        else:
            usssa_db[team_name]['usssa_class'] = team_class
        
        if team_name not in teams_db.keys():
            create_team(team_name, team_class, team_age, team_city, team_state, team_coach, teams_db)
        else:
            teams_db[team_name]['usssa_class'] = team_class

        #team events
        team_link.click()

        #tournaments
        try:
            element = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'div > .table-responsive-sm > .table > tbody > tr'))
            )
        except TimeoutException:
            continue
    
        games = driver.find_elements_by_css_selector("div[ng-repeat='e in team.completedGames']")
        tournaments = driver.find_elements_by_css_selector("tr[ng-repeat='event in team.events | filter:keyword']")

        for t in tournaments:
            t_info = t.find_elements_by_css_selector('td')
            t_placement = t_info.pop(0).text.strip()
            t_date = t_info.pop(0).text.strip()
            t_name = t_info.pop(1).text.strip()

            tournament = {
                't_id': "", 't_start_date': t_date, "t_sanction": 'USSSA',
                't_placement': t_placement, 't_name': t_name, 't_location': "", "t_director": ""
            }

            teams_db[team_name]['tournaments']['USSSA']['tournaments'][t_name] = tournament
            usssa_db[team_name]['tournaments']['USSSA']['tournaments'][t_name] = tournament

        for g in games:
            #it's a game
            bracket_division = g.find_element_by_css_selector('.row > div > a').get_attribute('href')
            tournament_name = g.find_element_by_css_selector('.row > div > a').text.strip()
            tournament_info = g.find_elements_by_css_selector('.row > div')
            tournament_location = tournament_info.pop(3).text.strip()
            tournament_location = tournament_location.replace('Location:', '')

            game_info = g.find_elements_by_css_selector('.table-responsive-sm > table > tbody > tr')
            game_data = ""
            for info in game_info:

                #its a game
                if info.get_attribute('ng-repeat-start') == 'game in e.games | filter:keyword':
                    game_info = info.find_elements_by_css_selector('td')
                    game_date = game_info.pop(0).text.strip()
                    team_1_score = game_info.pop(0).text.strip()
                    team_1_name = game_info.pop(1).text.strip()
                    team_1_state = game_info.pop(1).text.strip()
                    team_1_age = game_info.pop(1).text.strip()
                    team_1_result = ""

                #it's the same game
                elif info.get_attribute('ng-repeat-end') == '':
                    game_info = info.find_elements_by_css_selector('td')
                    team_2_score = game_info.pop(0).text.strip()
                    team_2_name = game_info.pop(1).text.strip()
                    team_2_state = game_info.pop(1).text.strip()
                    team_2_age = game_info.pop(1).text.strip()
                    team_2_result = ""
                    

                    (team_1_result, team_2_result) = get_results(int(team_1_score), int(team_2_score))

                    game_1 = {
                        "game_id" : str(game_id) + team_1_result, "game_date":game_date, "team_age": team_1_age, 
                        "team_name": team_1_name, "score": int(team_1_score), "tournament_santion": "USSSA", 
                        "bracket_divison": bracket_division, "tournament_name": tournament_name,
                        "tournament_location": tournament_location,
                    }

                    game_2 = {
                        "game_id" : str(game_id) + team_2_result, "game_date":game_date, "team_age": team_2_age, 
                        "team_name": team_2_name, "score": int(team_2_score), "tournament_santion": "USSSA", 
                        "bracket_divison": bracket_division, "tournament_name": tournament_name,
                        "tournament_location": tournament_location, 
                    }

                    if tournament_name in usssa_db[team_name]['tournaments']['USSSA']['tournaments'].keys():
                        teams_db[team_name]['tournaments']['USSSA']['tournaments'][t_name]['t_location'] = tournament_location    
                        usssa_db[team_name]['tournaments']['USSSA']['tournaments'][t_name]['t_location'] = tournament_location    

                    teams_db[team_name]['tournaments']['USSSA']['games'][str(game_id) + team_1_result] = game_1
                    teams_db[team_name]['tournaments']['USSSA']['games'][str(game_id) + team_2_result] = game_2
                    usssa_db[team_name]['tournaments']['USSSA']['games'][str(game_id) + team_1_result] = game_1
                    usssa_db[team_name]['tournaments']['USSSA']['games'][str(game_id) + team_2_result] = game_2

                    game_id += 1
        count += 1

    driver.quit()
    return game_id

def main(game_id, teams_db, usssa_db):
    wb = load_workbook('usssa_teams.xlsx')
    wb_sheet = wb['Sheet1']

    for row in wb_sheet.iter_rows():
        team_class = row[0].value
        team_url = row[1].value
        game_id = getTeamInfo(usssa_db, teams_db, team_class, team_url, game_id)

links = {}
teams_db = {}
usssa_db = {}
game_id = 8000

with open('./data_out/fasa_usfa_data.json', 'r') as data:
    teams_db = json.load(data)

data.close()

main(game_id, teams_db, usssa_db)

print("USSA Almost Finished!")

with open('./data_out/fasa_usfa_usssa_data.json', 'w') as fud:
    json.dump(teams_db, fud)

fud.close()

with open('./data_out/usssa_data.json', 'w') as ud:
    json.dump(usssa_db, ud)

ud.close()

print("Finished!")