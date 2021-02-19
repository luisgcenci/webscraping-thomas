from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time
import json
from openpyxl import Workbook, load_workbook

PATH = "../chromedriver.exe"

def get_coaches(teams_db, team_url):
    driver = webdriver.Chrome(PATH)
    driver.get(team_url)

    teams = driver.find_elements_by_css_selector('.table-responsive-sm > table > tbody > tr')
    len_teams = len(teams)
    
    for i in range(len_teams):

        try:
            team = driver.find_elements_by_css_selector('.table-responsive-sm > table > tbody > tr').pop(i)
            team_name = team.find_element_by_tag_name('td').text.strip()
            
            team_info = team.find_elements_by_tag_name('td')
            team_coach = team_info.pop(3).text.strip()

            teams_db[team_name] = {
                'team_name': team_name,
                'coach': team_coach
            }

        except (NoSuchElementException, IndexError) as e:
            break
    
    driver.quit()

def main(teams_db):
    wb = load_workbook('usssa_teams.xlsx')
    wb_sheet = wb['Sheet1']
    count = 0
    for row in wb_sheet.iter_rows():
        print(count)
        team_class = row[0].value
        team_url = row[1].value
        get_coaches(teams_db, team_url)
        count += 1

teams_db = {}
main(teams_db)

with open('usssa_coach_list.json', 'w') as file:
    json.dump(teams_db, file)

